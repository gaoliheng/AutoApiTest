import json
import re
from datetime import datetime
from typing import Optional
from dataclasses import dataclass, field, asdict

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QClipboard
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTextEdit,
    QComboBox,
    QGroupBox,
    QFileDialog,
    QMessageBox,
    QProgressBar,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QLineEdit,
    QDialog,
    QCheckBox,
    QApplication,
    QInputDialog,
    QDialogButtonBox,
    QScrollArea,
    QFrame,
)

from models.ai_model import AIModel
from models.test_case_history import TestCaseHistory
from ai.client import AIClient, AIModelConfig, ChatMessage, MessageRole
from core.export_service import ExportService, ExportFormat
from ui.styles import style_manager
from ui.pages.base_page import BasePage
from ui.dialogs.test_case_history_dialog import TestCaseHistoryDialog
from ui.dialogs.dimension_config_dialog import DimensionConfigDialog
from utils.logger import get_logger
from utils.config import config

_logger = get_logger("ui.test_case_page")


@dataclass
class TestCaseData:
    name: str = ""
    method: str = "GET"
    params: str = ""
    body: str = ""
    expected_status: int = 200
    assertions: str = ""
    
    def to_dict(self) -> dict:
        result = asdict(self)
        try:
            if self.params.strip():
                result["params"] = json.loads(self.params)
            else:
                result["params"] = {}
        except:
            result["params"] = {}
        
        try:
            if self.body.strip():
                result["body"] = json.loads(self.body)
            else:
                result["body"] = {}
        except:
            result["body"] = {}
        
        return result


class TextEditDialog(QDialog):
    def __init__(self, title: str, content: str, placeholder: str = "", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(500, 350)
        self._result = content
        
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        self._text_edit = QTextEdit()
        self._text_edit.setPlainText(content)
        self._text_edit.setPlaceholderText(placeholder)
        self._text_edit.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 8px;
                font-size: 13px;
                background: #fafafa;
            }
        """)
        layout.addWidget(self._text_edit)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        style_manager.apply_style(cancel_btn, "button_secondary")
        btn_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("确定")
        ok_btn.setFixedWidth(80)
        ok_btn.clicked.connect(self._on_ok)
        style_manager.apply_style(ok_btn, "button_primary")
        btn_layout.addWidget(ok_btn)
        
        layout.addLayout(btn_layout)
    
    def _on_ok(self) -> None:
        self._result = self._text_edit.toPlainText()
        self.accept()
    
    def get_result(self) -> str:
        return self._result


class GenerateThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(list, str, str, str)
    error = pyqtSignal(str)

    def __init__(
        self,
        ai_model: AIModel,
        doc_content: str,
        dimensions: list[dict],
        parent: Optional[QWidget] = None
    ) -> None:
        super().__init__(parent)
        self._ai_model = ai_model
        self._doc_content = doc_content
        self._dimensions = dimensions

    def run(self) -> None:
        try:
            self.progress.emit("正在初始化 AI 模型...")

            config = AIModelConfig(
                name=self._ai_model.name,
                api_base_url=self._ai_model.api_base,
                api_key=self._ai_model.api_key,
                model_name=self._ai_model.model_name,
                timeout=300.0,
            )

            client = AIClient(config)

            self.progress.emit("正在分析接口文档...")

            dimension_text = self._build_dimension_text()

            system_prompt = f"""你是一个专业的API测试工程师。你的任务是根据用户提供的API接口文档生成测试用例。

请按照以下JSON格式输出测试用例列表：
```json
{{
    "base_url": "http://api.example.com:8080",
    "api_path": "/api/v1/users",
    "common_headers": {{
        "Authorization": "Bearer your_token_here",
        "Content-Type": "application/json"
    }},
    "test_cases": [
        {{
            "name": "测试用例名称",
            "method": "GET/POST/PUT/DELETE",
            "params": "page=1&size=10 或 JSON格式",
            "body": "请求体JSON字符串",
            "expected_status": 200,
            "assertions": "断言描述，如：响应状态码为200，data.list长度大于0，msg等于成功"
        }}
    ]
}}
```

【重要】本次需要生成的测试维度：
{dimension_text}

【强制要求】必须严格按照上述维度生成测试用例：
1. 为每个启用的维度生成至少 1-2 个测试用例
2. 每个测试用例的 name 或 description 必须明确说明它属于哪个维度
3. 例如：如果启用了"异常场景测试"，必须生成参数缺失、类型错误等异常测试用例
4. 不要生成与指定维度无关的测试用例

请确保：
1. 从文档中提取 base_url（包含协议、IP、端口），如果没有找到则设为空字符串
2. 从文档中提取 api_path（接口路径），如果没有找到则设为空字符串
3. 【重要】从文档中提取 common_headers（通用请求头），特别关注：
   - Authorization：认证信息（如 Bearer token、API Key 等），这是用户最关注的
   - Content-Type：内容类型
   - 其他必要的请求头
   如果文档中提到需要认证但未给出具体token，Authorization 值使用占位符如 "Bearer your_token_here"
   如果文档中对请求头的描述是自然语言（如"需要在Header中传入token"），也可以提取为JSON格式
   如果没有找到任何请求头，则设为空对象 {{}}
5. params 和 body 使用自然语言或JSON格式描述
6. assertions 使用自然语言描述断言条件
7. 只输出JSON，不要包含其他说明文字"""

            user_message = f"以下是API接口文档：\n\n{self._doc_content}\n\n请根据以上接口文档生成测试用例。"

            messages = [
                ChatMessage(role=MessageRole.SYSTEM, content=system_prompt),
                ChatMessage(role=MessageRole.USER, content=user_message),
            ]

            self.progress.emit("正在生成测试用例...")

            response = client.chat(messages, temperature=0.7, max_tokens=16384)

            content = response["choices"][0]["message"]["content"]

            test_cases, base_url, api_path, common_headers = self._parse_response(content)

            client.close()

            self.progress.emit("测试用例生成完成!")
            self.finished.emit(test_cases, base_url, api_path, common_headers)

        except Exception as e:
            self.error.emit(f"生成失败: {str(e)}")

    def _build_dimension_text(self) -> str:
        if not self._dimensions:
            _logger.warning("维度配置为空，使用默认提示")
            return "正常场景测试、异常场景测试"

        lines = []
        priority_markers = {"high": "⭐⭐⭐", "medium": "⭐⭐", "low": "⭐"}
        for i, dim in enumerate(self._dimensions, 1):
            name = dim.get("name", f"维度{i}")
            desc = dim.get("description", "")
            priority = dim.get("priority", "medium")
            marker = priority_markers.get(priority, "⭐")
            lines.append(f"{i}. {name}（{marker}）：{desc}")
        result = "\n".join(lines)
        _logger.debug(f"生成的维度配置文本: {result}")
        return result
    
    def _parse_response(self, content: str) -> tuple[list[dict], str, str, str]:
        json_str = content
        if "```json" in content:
            start = content.find("```json") + 7
            end = content.find("```", start)
            json_str = content[start:end].strip()
        elif "```" in content:
            start = content.find("```") + 3
            end = content.find("```", start)
            json_str = content[start:end].strip()
        
        data = json.loads(json_str)
        
        if isinstance(data, list):
            return data, "", "", ""
        
        base_url = data.get("base_url", "")
        api_path = data.get("api_path", "")
        common_headers_dict = data.get("common_headers", {})
        test_cases = data.get("test_cases", [])
        
        if not isinstance(test_cases, list):
            test_cases = [test_cases]
        
        common_headers = ""
        if isinstance(common_headers_dict, dict) and common_headers_dict:
            common_headers = json.dumps(common_headers_dict, ensure_ascii=False, indent=2)
        
        return test_cases, base_url, api_path, common_headers


_test_cases_store: list[TestCaseData] = []
_base_url_store: str = ""
_api_path_store: str = ""
_common_headers_store: str = ""
_selected_indices_store: list[int] = []
_api_doc_store: str = ""


def get_test_cases() -> list[TestCaseData]:
    return _test_cases_store


def set_test_cases(cases: list[TestCaseData]) -> None:
    global _test_cases_store
    _test_cases_store = cases


def get_base_url() -> str:
    return _base_url_store


def set_base_url(url: str) -> None:
    global _base_url_store
    _base_url_store = url


def get_api_path() -> str:
    return _api_path_store


def set_api_path(path: str) -> None:
    global _api_path_store
    _api_path_store = path


def get_common_headers() -> str:
    return _common_headers_store


def set_common_headers(headers: str) -> None:
    global _common_headers_store
    _common_headers_store = headers


def get_selected_indices() -> list[int]:
    return _selected_indices_store


def set_selected_indices(indices: list[int]) -> None:
    global _selected_indices_store
    _selected_indices_store = indices


def get_api_doc() -> str:
    return _api_doc_store


def set_api_doc(doc: str) -> None:
    global _api_doc_store
    _api_doc_store = doc


HTTP_STATUS_CODES = {
    "200": "200 OK - 请求成功",
    "201": "201 Created - 创建成功",
    "204": "204 No Content - 无内容",
    "400": "400 Bad Request - 请求参数错误",
    "401": "401 Unauthorized - 未授权",
    "403": "403 Forbidden - 禁止访问",
    "404": "404 Not Found - 资源不存在",
    "500": "500 Internal Server Error - 服务器错误",
    "502": "502 Bad Gateway - 网关错误",
    "503": "503 Service Unavailable - 服务不可用",
}


class TestCasePage(BasePage):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__("测试用例", parent)
        self._generate_thread: Optional[GenerateThread] = None
        self._init_content()
        self._load_ai_models()

    def _init_content(self) -> None:
        self._main_bar = QWidget()
        self._main_bar.setObjectName("main_bar")
        self._main_bar.setStyleSheet("""
            QWidget {
                background-color: #fafafa;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                padding: 12px 15px;
            }
        """)
        main_layout = QHBoxLayout(self._main_bar)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(15, 12, 15, 12)

        input_section = QWidget()
        input_layout = QHBoxLayout(input_section)
        input_layout.setContentsMargins(0, 0, 0, 0)
        input_layout.setSpacing(8)

        self._edit_doc_btn = QPushButton("编辑文档")
        self._edit_doc_btn.clicked.connect(self._show_doc_dialog)
        self._edit_doc_btn.setStyleSheet("""
            QPushButton {
                background-color: #e8f5e9;
                color: #388e3c;
                border: 1px solid #a5d6a7;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #c8e6c9;
            }
        """)
        input_layout.addWidget(self._edit_doc_btn)

        self._upload_btn = QPushButton("上传文件")
        self._upload_btn.clicked.connect(self._upload_file)
        self._upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #fff3e0;
                color: #e65100;
                border: 1px solid #ffcc80;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #ffe0b2;
            }
        """)
        input_layout.addWidget(self._upload_btn)

        main_layout.addWidget(input_section)

        arrow_label = QLabel("➜")
        arrow_label.setStyleSheet("color: #7b1fa2; font-size: 18px; font-weight: bold; margin: 0 10px;")
        main_layout.addWidget(arrow_label)

        generate_section = QWidget()
        generate_layout = QHBoxLayout(generate_section)
        generate_layout.setContentsMargins(0, 0, 0, 0)
        generate_layout.setSpacing(8)

        self._model_combo = QComboBox()
        self._model_combo.setMinimumWidth(100)
        self._model_combo.setStyleSheet("""
            QComboBox {
                background-color: #ffffff;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px 10px;
                font-size: 13px;
            }
        """)
        generate_layout.addWidget(self._model_combo)

        self._dimension_config_btn = QPushButton("配置维度")
        self._dimension_config_btn.clicked.connect(self._show_dimension_config)
        self._dimension_config_btn.setStyleSheet("""
            QPushButton {
                background-color: #f3e5f5;
                color: #7b1fa2;
                border: 1px solid #ce93d8;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #e1bee7;
            }
        """)
        generate_layout.addWidget(self._dimension_config_btn)

        self._generate_btn = QPushButton("生成测试用例")
        self._generate_btn.clicked.connect(self._generate_test_cases)
        self._generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #7b1fa2;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 24px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #9c27b0;
            }
        """)
        generate_layout.addWidget(self._generate_btn)

        main_layout.addWidget(generate_section)

        arrow_label2 = QLabel("➜")
        arrow_label2.setStyleSheet("color: #7b1fa2; font-size: 18px; font-weight: bold; margin: 0 10px;")
        main_layout.addWidget(arrow_label2)

        manage_section = QWidget()
        manage_layout = QHBoxLayout(manage_section)
        manage_layout.setContentsMargins(0, 0, 0, 0)
        manage_layout.setSpacing(8)

        self._save_history_btn = QPushButton("保存历史")
        self._save_history_btn.clicked.connect(self._save_to_history)
        self._save_history_btn.setStyleSheet("""
            QPushButton {
                background-color: #e3f2fd;
                color: #1565c0;
                border: 1px solid #90caf9;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #bbdefb;
            }
        """)
        manage_layout.addWidget(self._save_history_btn)

        self._history_btn = QPushButton("历史记录")
        self._history_btn.clicked.connect(self._show_history_dialog)
        self._history_btn.setStyleSheet("""
            QPushButton {
                background-color: #f5f5f5;
                color: #424242;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #eeeeee;
            }
        """)
        manage_layout.addWidget(self._history_btn)

        main_layout.addWidget(manage_section)
        main_layout.addStretch()

        self.add_widget(self._main_bar)

        self._base_url_edit = QLineEdit()
        self._base_url_edit.setVisible(False)
        self._base_url_edit.textChanged.connect(self._save_config)
        self.add_widget(self._base_url_edit)

        self._api_path_edit = QLineEdit()
        self._api_path_edit.setVisible(False)
        self._api_path_edit.textChanged.connect(self._save_config)
        self.add_widget(self._api_path_edit)

        self._api_info_bar = QWidget()
        self._api_info_bar.setVisible(False)
        self._api_info_bar.setStyleSheet("""
            QWidget {
                background-color: #e3f2fd;
                border: 1px solid #90caf9;
                border-radius: 6px;
                padding: 8px 12px;
            }
        """)
        api_info_layout = QHBoxLayout(self._api_info_bar)
        api_info_layout.setContentsMargins(10, 5, 10, 5)
        api_info_layout.setSpacing(10)

        base_url_label = QLabel("Base URL:")
        base_url_label.setStyleSheet("font-size: 12px; color: #1565c0; font-weight: bold;")
        api_info_layout.addWidget(base_url_label)

        self._base_url_display = QLineEdit()
        self._base_url_display.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #90caf9;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
            }
        """)
        self._base_url_display.textChanged.connect(self._on_base_url_changed)
        api_info_layout.addWidget(self._base_url_display, 1)

        api_path_label = QLabel("路径:")
        api_path_label.setStyleSheet("font-size: 12px; color: #1565c0; font-weight: bold;")
        api_info_layout.addWidget(api_path_label)

        self._api_path_display = QLineEdit()
        self._api_path_display.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #90caf9;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
            }
        """)
        self._api_path_display.textChanged.connect(self._on_api_path_changed)
        api_info_layout.addWidget(self._api_path_display, 1)

        self.add_widget(self._api_info_bar)

        progress_layout = QHBoxLayout()
        progress_layout.setContentsMargins(0, 8, 0, 0)
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, 0)
        self._progress_bar.setVisible(False)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setStyleSheet("""
            QProgressBar {
                background-color: #f3e5f5;
                border: none;
                border-radius: 9px;
                text-align: center;
                font-size: 12px;
                color: #7b1fa2;
            }
            QProgressBar::chunk {
                background-color: #9c27b0;
                border-radius: 9px;
            }
        """)
        progress_layout.addWidget(self._progress_bar)

        self._status_label = QLabel("")
        self._status_label.setStyleSheet("color: #757575; font-size: 12px; margin-left: 10px;")
        progress_layout.addWidget(self._status_label)
        progress_layout.addStretch()
        self.add_layout(progress_layout)

        table_group = QGroupBox("测试用例列表")
        table_group.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                font-size: 14px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 10px;
                color: #1565c0;
            }
        """)
        table_layout = QVBoxLayout(table_group)
        table_layout.setSpacing(10)
        table_layout.setContentsMargins(15, 15, 15, 15)

        table_btn_layout = QHBoxLayout()
        table_btn_layout.setSpacing(8)

        self._select_all_cb = QCheckBox("全选")
        self._select_all_cb.stateChanged.connect(self._on_select_all)
        self._select_all_cb.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-size: 13px;
                color: #424242;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 4px;
                border: 2px solid #bdbdbd;
            }
            QCheckBox::indicator:checked {
                background-color: #2196f3;
                border-color: #2196f3;
            }
        """)
        table_btn_layout.addWidget(self._select_all_cb)

        self._add_row_btn = QPushButton("添加行")
        self._add_row_btn.clicked.connect(self._add_row)
        self._add_row_btn.setStyleSheet("""
            QPushButton {
                background-color: #e3f2fd;
                color: #1565c0;
                border: 1px solid #90caf9;
                border-radius: 4px;
                padding: 5px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #bbdefb;
            }
        """)
        table_btn_layout.addWidget(self._add_row_btn)

        self._delete_row_btn = QPushButton("删除")
        self._delete_row_btn.clicked.connect(self._delete_selected_rows)
        self._delete_row_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffebee;
                color: #c62828;
                border: 1px solid #ef9a9a;
                border-radius: 4px;
                padding: 5px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #ffcdd2;
            }
        """)
        table_btn_layout.addWidget(self._delete_row_btn)

        self._clear_table_btn = QPushButton("清空")
        self._clear_table_btn.clicked.connect(self._clear_table)
        self._clear_table_btn.setStyleSheet("""
            QPushButton {
                background-color: #fff3e0;
                color: #e65100;
                border: 1px solid #ffcc80;
                border-radius: 4px;
                padding: 5px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #ffe0b2;
            }
        """)
        table_btn_layout.addWidget(self._clear_table_btn)

        self._export_btn = QPushButton("导出")
        self._export_btn.clicked.connect(self._export_test_cases)
        self._export_btn.setStyleSheet("""
            QPushButton {
                background-color: #e8f5e9;
                color: #388e3c;
                border: 1px solid #a5d6a7;
                border-radius: 4px;
                padding: 5px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c8e6c9;
            }
        """)
        table_btn_layout.addWidget(self._export_btn)

        table_btn_layout.addStretch()

        self._count_label = QLabel("共 0 条")
        self._count_label.setStyleSheet("color: #757575; font-size: 12px; font-weight: 500;")
        table_btn_layout.addWidget(self._count_label)

        self._generate_script_btn = QPushButton("去生成脚本")
        self._generate_script_btn.clicked.connect(self._go_to_script_page)
        self._generate_script_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565c0;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976d2;
            }
        """)
        table_btn_layout.addWidget(self._generate_script_btn)

        table_layout.addLayout(table_btn_layout)

        self._table = QTableWidget()
        self._table.setColumnCount(7)
        self._table.setHorizontalHeaderLabels(["勾选", "用例名称", "方法", "请求参数", "请求体", "HTTP状态码", "断言"])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(0, 60)
        self._table.setColumnWidth(1, 200)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 200)
        self._table.setColumnWidth(4, 200)
        self._table.setColumnWidth(5, 180)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setDefaultSectionSize(40)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setMinimumSectionSize(50)
        self._table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        self._table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                background-color: #ffffff;
                gridline-color: #f5f5f5;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 8px 12px;
                border-bottom: 1px solid #f5f5f5;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: #1565c0;
            }
            QTableWidget::item:hover {
                background-color: #f5f5f5;
            }
            QHeaderView::section {
                background-color: #fafafa;
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid #e0e0e0;
                font-weight: 600;
                font-size: 13px;
                color: #424242;
            }
            QHeaderView::section:hover {
                background-color: #f5f5f5;
            }
        """)
        table_layout.addWidget(self._table, 1)

        self.add_widget(table_group, 1)

        self._doc_input = QTextEdit()
        self._doc_input.setVisible(False)
        self._doc_input.textChanged.connect(self._on_doc_changed)

    def _show_doc_dialog(self) -> None:
        current_doc = self._doc_input.toPlainText()
        dialog = TextEditDialog(
            "编辑接口文档",
            current_doc,
            "请在此输入接口文档内容，支持 Markdown、JSON、YAML 等格式...",
            self
        )
        dialog.setMinimumSize(700, 500)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._doc_input.setPlainText(dialog.get_result())
            self._update_doc_status()

    def _clear_doc(self) -> None:
        self._doc_input.clear()
        self._update_doc_status()

    def _show_config_panel(self) -> None:
        self._api_info_bar.setVisible(False)

    def _on_base_url_changed(self, text: str) -> None:
        self._base_url_edit.setText(text)

    def _on_api_path_changed(self, text: str) -> None:
        self._api_path_edit.setText(text)

    def _on_doc_changed(self) -> None:
        self._update_doc_status()

    def _update_doc_status(self) -> None:
        doc = self._doc_input.toPlainText().strip()
        set_api_doc(doc)
        if doc:
            self._edit_doc_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4caf50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background-color: #66bb6a;
                }
            """)
        else:
            self._edit_doc_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e8f5e9;
                    color: #388e3c;
                    border: 1px solid #a5d6a7;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background-color: #c8e6c9;
                }
            """)
    
    def _on_select_all(self, state: int) -> None:
        checked = state == Qt.CheckState.Checked.value
        for row in range(self._table.rowCount()):
            checkbox = self._table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(checked)
    
    def _get_checked_rows(self) -> list[int]:
        checked_rows = []
        for row in range(self._table.rowCount()):
            checkbox = self._table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                checked_rows.append(row)
        return checked_rows
    
    def _on_cell_double_clicked(self, row: int, column: int) -> None:
        if column == 3:
            current_text = self._table.item(row, column).text() if self._table.item(row, column) else ""
            dialog = TextEditDialog(
                "编辑请求参数",
                current_text,
                "请输入请求参数，支持自然语言或JSON格式：\n\n"
                "示例1（自然语言）：page=1, size=10, name包含'张三'\n\n"
                "示例2（JSON）：{\"page\": 1, \"size\": 10}",
                self
            )
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self._table.setItem(row, column, QTableWidgetItem(dialog.get_result()))
                self._save_to_store()
        
        elif column == 4:
            current_text = self._table.item(row, column).text() if self._table.item(row, column) else ""
            dialog = TextEditDialog(
                "编辑请求体",
                current_text,
                "请输入请求体，支持自然语言或JSON格式：\n\n"
                "示例1（自然语言）：用户名为test，密码为123456\n\n"
                "示例2（JSON）：{\"username\": \"test\", \"password\": \"123456\"}",
                self
            )
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self._table.setItem(row, column, QTableWidgetItem(dialog.get_result()))
                self._save_to_store()
        
        elif column == 6:
            current_text = self._table.item(row, column).text() if self._table.item(row, column) else ""
            dialog = TextEditDialog(
                "编辑断言",
                current_text,
                "请输入断言条件（自然语言）：\n\n"
                "示例：\n"
                "- 响应状态码为200\n"
                "- data.list长度大于0\n"
                "- msg等于'成功'\n"
                "- code为0",
                self
            )
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self._table.setItem(row, column, QTableWidgetItem(dialog.get_result()))
                self._save_to_store()
    
    def _go_to_script_page(self) -> None:
        if self._table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "请先添加或生成测试用例")
            return
        
        base_url = self._base_url_edit.text().strip()
        if not base_url:
            QMessageBox.warning(self, "提示", "请先配置 Base URL")
            return
        
        api_path = self._api_path_edit.text().strip()
        if not api_path:
            QMessageBox.warning(self, "提示", "请先配置接口路径")
            return
        
        checked_rows = self._get_checked_rows()
        if checked_rows:
            set_selected_indices(checked_rows)
        else:
            set_selected_indices(list(range(self._table.rowCount())))
        
        self._save_to_store()
        
        from ui.main_window import MainWindow
        window = self.window()
        if isinstance(window, MainWindow):
            window._nav_list.setCurrentRow(3)
    
    def _load_ai_models(self) -> None:
        self._model_combo.clear()
        models = AIModel.get_all()
        
        for model in models:
            default_marker = " (默认)" if model.is_default else ""
            self._model_combo.addItem(f"{model.name}{default_marker}", model.id)
        
        default_model = AIModel.get_default()
        if default_model:
            index = self._model_combo.findData(default_model.id)
            if index >= 0:
                self._model_combo.setCurrentIndex(index)
    
    def _upload_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Markdown 文件",
            "",
            "Markdown Files (*.md);;All Files (*)"
        )
        
        if file_path:
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
                self._doc_input.setText(content)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件失败: {str(e)}")
    
    def _clear_input(self) -> None:
        self._doc_input.clear()
        self._status_label.clear()
    
    def _save_config(self) -> None:
        set_base_url(self._base_url_edit.text().strip())
        set_api_path(self._api_path_edit.text().strip())
        set_common_headers("")
    
    def _generate_test_cases(self) -> None:
        doc_content = self._doc_input.toPlainText().strip()
        if not doc_content:
            QMessageBox.warning(self, "提示", "请输入接口文档内容")
            return

        set_api_doc(doc_content)

        model_id = self._model_combo.currentData()
        if not model_id:
            QMessageBox.warning(self, "提示", "请选择 AI 模型")
            return

        ai_model = AIModel.get_by_id(model_id)
        if not ai_model:
            QMessageBox.warning(self, "提示", "AI 模型不存在")
            return

        enabled_dimensions = config.get_enabled_dimensions()
        _logger.info(f"开始生成测试用例: model={ai_model.name}, doc_length={len(doc_content)}, dimensions_count={len(enabled_dimensions)}, dimensions={enabled_dimensions}")

        self._generate_btn.setEnabled(False)
        self._progress_bar.setVisible(True)
        self._status_label.setText("正在生成测试用例...")

        self._generate_thread = GenerateThread(
            ai_model=ai_model,
            doc_content=doc_content,
            dimensions=enabled_dimensions,
        )
        self._generate_thread.progress.connect(self._on_progress)
        self._generate_thread.finished.connect(self._on_finished)
        self._generate_thread.error.connect(self._on_error)
        self._generate_thread.start()
    
    def _on_progress(self, message: str) -> None:
        self._status_label.setText(message)
        _logger.debug(f"生成进度: {message}")
    
    def _on_finished(self, test_cases: list[dict], base_url: str, api_path: str, common_headers: str) -> None:
        self._generate_btn.setEnabled(True)
        self._progress_bar.setVisible(False)
        self._status_label.setText(f"成功生成 {len(test_cases)} 个测试用例")

        _logger.info(f"测试用例生成完成: count={len(test_cases)}, base_url={base_url}, api_path={api_path}, headers={common_headers}")

        if base_url or api_path:
            self._base_url_edit.setText(base_url)
            self._api_path_edit.setText(api_path)
            self._base_url_display.setText(base_url)
            self._api_path_display.setText(api_path)
            self._api_info_bar.setVisible(True)

        self._load_cases_to_table(test_cases)
        self._save_config()
    
    def _on_error(self, error_message: str) -> None:
        self._generate_btn.setEnabled(True)
        self._progress_bar.setVisible(False)
        self._status_label.setText("生成失败")
        _logger.error(f"测试用例生成失败: {error_message}")
        QMessageBox.critical(self, "错误", error_message)
    
    def _load_cases_to_table(self, cases: list[dict]) -> None:
        self._table.setRowCount(len(cases))
        
        for row, case_data in enumerate(cases):
            name = case_data.get("name", "")
            method = case_data.get("method", "GET")
            params = case_data.get("params", "")
            body = case_data.get("body", "")
            expected_status = case_data.get("expected_status", 200)
            assertions = case_data.get("assertions", "")
            
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 0px;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                    border-radius: 3px;
                    border: 2px solid #bdbdbd;
                }
                QCheckBox::indicator:checked {
                    background-color: #2196f3;
                    border-color: #2196f3;
                }
                QCheckBox::indicator:hover {
                    border-color: #2196f3;
                }
            """)
            self._table.setCellWidget(row, 0, checkbox)
            
            self._table.setItem(row, 1, QTableWidgetItem(name))
            self._table.setItem(row, 2, QTableWidgetItem(method))
            self._table.setItem(row, 3, QTableWidgetItem(str(params) if params else ""))
            self._table.setItem(row, 4, QTableWidgetItem(str(body) if body else ""))
            
            status_combo = QComboBox()
            for code, desc in HTTP_STATUS_CODES.items():
                status_combo.addItem(desc, code)
            status_combo.setCurrentText(HTTP_STATUS_CODES.get(str(expected_status), str(expected_status)))
            status_combo.currentIndexChanged.connect(lambda _, r=row: self._on_status_changed(r))
            status_combo.setStyleSheet("""
                QComboBox {
                    background-color: #fafafa;
                    border: 1px solid #e0e0e0;
                    border-radius: 4px;
                    padding: 4px 8px;
                    font-size: 12px;
                    min-width: 140px;
                }
                QComboBox:hover {
                    border-color: #bdbdbd;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox QAbstractItemView {
                    background-color: #ffffff;
                    border: 1px solid #e0e0e0;
                    selection-background-color: #e3f2fd;
                    min-width: 200px;
                }
            """)
            self._table.setCellWidget(row, 5, status_combo)
            
            self._table.setItem(row, 6, QTableWidgetItem(str(assertions) if assertions else ""))
        
        self._select_all_cb.setChecked(False)
        self._update_count()
        self._save_to_store()
    
    def _on_status_changed(self, row: int) -> None:
        self._save_to_store()
    
    def _add_row(self) -> None:
        row = self._table.rowCount()
        self._table.insertRow(row)
        
        checkbox = QCheckBox()
        checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 0px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border-radius: 3px;
                border: 2px solid #bdbdbd;
            }
            QCheckBox::indicator:checked {
                background-color: #2196f3;
                border-color: #2196f3;
            }
            QCheckBox::indicator:hover {
                border-color: #2196f3;
            }
        """)
        self._table.setCellWidget(row, 0, checkbox)
        
        self._table.setItem(row, 1, QTableWidgetItem(""))
        self._table.setItem(row, 2, QTableWidgetItem("GET"))
        self._table.setItem(row, 3, QTableWidgetItem(""))
        self._table.setItem(row, 4, QTableWidgetItem(""))
        
        status_combo = QComboBox()
        for code, desc in HTTP_STATUS_CODES.items():
            status_combo.addItem(desc, code)
        status_combo.setCurrentIndex(0)
        status_combo.currentIndexChanged.connect(lambda _, r=row: self._on_status_changed(r))
        status_combo.setStyleSheet("""
            QComboBox {
                background-color: #fafafa;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
                min-width: 140px;
            }
            QComboBox:hover {
                border-color: #bdbdbd;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                border: 1px solid #e0e0e0;
                selection-background-color: #e3f2fd;
                min-width: 200px;
            }
        """)
        self._table.setCellWidget(row, 5, status_combo)
        
        self._table.setItem(row, 6, QTableWidgetItem(""))
        self._update_count()
        self._save_to_store()
    
    def _delete_selected_rows(self) -> None:
        checked_rows = self._get_checked_rows()
        
        if not checked_rows:
            QMessageBox.warning(self, "提示", "请先勾选要删除的行")
            return
        
        reply = QMessageBox.question(
            self,
            "确认",
            f"确定要删除选中的 {len(checked_rows)} 条测试用例吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            for row in sorted(checked_rows, reverse=True):
                self._table.removeRow(row)
            
            self._select_all_cb.setChecked(False)
            self._update_count()
            self._save_to_store()
    
    def _clear_table(self) -> None:
        if self._table.rowCount() == 0:
            return
        
        reply = QMessageBox.question(
            self,
            "确认",
            "确定要清空所有测试用例吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self._table.setRowCount(0)
            self._select_all_cb.setChecked(False)
            self._update_count()
            self._save_to_store()
    
    def _update_count(self) -> None:
        count = self._table.rowCount()
        self._count_label.setText(f"共 {count} 条")
    
    def _save_to_store(self) -> None:
        cases = []
        for row in range(self._table.rowCount()):
            status_widget = self._table.cellWidget(row, 5)
            status_code = status_widget.currentData() if status_widget else "200"
            
            case = TestCaseData(
                name=self._table.item(row, 1).text() if self._table.item(row, 1) else "",
                method=self._table.item(row, 2).text() if self._table.item(row, 2) else "GET",
                params=self._table.item(row, 3).text() if self._table.item(row, 3) else "",
                body=self._table.item(row, 4).text() if self._table.item(row, 4) else "",
                expected_status=int(status_code),
                assertions=self._table.item(row, 6).text() if self._table.item(row, 6) else "",
            )
            cases.append(case)
        
        set_test_cases(cases)
        self._save_config()
    
    def _export_test_cases(self) -> None:
        if self._table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "表格中没有测试用例")
            return
        
        checked_rows = self._get_checked_rows()
        
        if not checked_rows:
            reply = QMessageBox.question(
                self,
                "确认导出",
                "没有勾选任何测试用例，是否导出所有测试用例？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            rows_to_export = range(self._table.rowCount())
        else:
            rows_to_export = checked_rows
        
        format_dialog = QInputDialog(self)
        format_dialog.setWindowTitle("选择导出格式")
        format_dialog.setLabelText("请选择导出格式:")
        format_dialog.setComboBoxItems(["Excel (.xlsx)", "JSON (.json)", "YAML (.yaml)"])
        format_dialog.setTextValue("Excel (.xlsx)")
        
        if format_dialog.exec() != QDialog.DialogCode.Accepted:
            return
        
        selected_format = format_dialog.textValue()
        
        if "Excel" in selected_format:
            export_format = ExportFormat.EXCEL
            file_filter = "Excel Files (*.xlsx);;All Files (*)"
            default_ext = ".xlsx"
        elif "JSON" in selected_format:
            export_format = ExportFormat.JSON
            file_filter = "JSON Files (*.json);;All Files (*)"
            default_ext = ".json"
        else:
            export_format = ExportFormat.YAML
            file_filter = "YAML Files (*.yaml);;All Files (*)"
            default_ext = ".yaml"
        
        from pathlib import Path
        from datetime import datetime
        import json
        import yaml
        
        default_dir = Path.home()
        if not default_dir.exists():
            default_dir = Path(".")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"test_cases_{timestamp}{default_ext}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存测试用例",
            str(default_dir / default_filename),
            file_filter
        )
        
        if not file_path:
            return
        
        try:
            test_cases_data = []
            for row in rows_to_export:
                status_widget = self._table.cellWidget(row, 5)
                status_code = status_widget.currentData() if status_widget else "200"
                
                tc_data = {
                    "id": row + 1,
                    "name": self._table.item(row, 1).text() if self._table.item(row, 1) else "",
                    "api_path": get_api_path(),
                    "method": self._table.item(row, 2).text() if self._table.item(row, 2) else "GET",
                    "headers": {},
                    "params": self._table.item(row, 3).text() if self._table.item(row, 3) else "",
                    "body": self._table.item(row, 4).text() if self._table.item(row, 4) else "",
                    "expected_status": int(status_code),
                    "assertions": self._table.item(row, 6).text() if self._table.item(row, 6) else "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat()
                }
                test_cases_data.append(tc_data)
            
            if export_format == ExportFormat.EXCEL:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                wb = Workbook()
                ws = wb.active
                ws.title = "测试用例"
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                headers = ["ID", "用例名称", "接口路径", "请求方法", "请求头", "请求参数", "请求体", "预期状态码", "断言", "创建时间"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                for row_num, tc in enumerate(test_cases_data, 2):
                    ws.cell(row=row_num, column=1, value=tc["id"])
                    ws.cell(row=row_num, column=2, value=tc["name"])
                    ws.cell(row=row_num, column=3, value=tc["api_path"])
                    ws.cell(row=row_num, column=4, value=tc["method"])
                    
                    headers_value = json.dumps(tc["headers"], ensure_ascii=False) if tc["headers"] else ""
                    ws.cell(row=row_num, column=5, value=headers_value)
                    
                    ws.cell(row=row_num, column=6, value=tc["params"])
                    
                    body_value = tc["body"] if tc["body"] else ""
                    ws.cell(row=row_num, column=7, value=body_value)
                    
                    ws.cell(row=row_num, column=8, value=tc["expected_status"])
                    
                    ws.cell(row=row_num, column=9, value=tc["assertions"])
                    
                    ws.cell(row=row_num, column=10, value=tc["created_at"][:19].replace("T", " "))
                    
                    for col_num in range(1, 11):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                column_widths = [8, 25, 35, 12, 30, 30, 30, 12, 35, 20]
                for col_num, width in enumerate(column_widths, 1):
                    ws.column_dimensions[chr(64 + col_num)].width = width
                
                wb.save(file_path)
                
            elif export_format == ExportFormat.JSON:
                export_data = {
                    "export_info": {
                        "exported_at": datetime.now().isoformat(),
                        "total_count": len(test_cases_data),
                        "version": "1.0"
                    },
                    "test_cases": test_cases_data
                }
                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                    
            else:
                export_data = {
                    "export_info": {
                        "exported_at": datetime.now().isoformat(),
                        "total_count": len(test_cases_data),
                        "version": "1.0"
                    },
                    "test_cases": test_cases_data
                }
                with open(file_path, "w", encoding="utf-8") as f:
                    yaml.dump(export_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            
            QMessageBox.information(self, "成功", f"测试用例已导出到: {file_path}")
            _logger.info(f"测试用例导出成功: {file_path}, 格式: {export_format.value}")
            
        except PermissionError:
            QMessageBox.critical(self, "错误", f"权限不足，无法保存到: {file_path}\n请选择其他目录或以管理员身份运行")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
            _logger.error(f"导出测试用例失败: {str(e)}")
    
    def refresh(self) -> None:
        self._load_ai_models()
    
    def _save_to_history(self) -> None:
        """保存当前用例到历史记录"""
        if self._table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "当前没有测试用例可保存")
            return
        
        # 获取当前用例数据
        test_cases = []
        for row in range(self._table.rowCount()):
            status_widget = self._table.cellWidget(row, 5)
            status_code = status_widget.currentData() if status_widget else "200"
            
            case = {
                "name": self._table.item(row, 1).text() if self._table.item(row, 1) else "",
                "method": self._table.item(row, 2).text() if self._table.item(row, 2) else "GET",
                "params": self._table.item(row, 3).text() if self._table.item(row, 3) else "",
                "body": self._table.item(row, 4).text() if self._table.item(row, 4) else "",
                "expected_status": int(status_code),
                "assertions": self._table.item(row, 6).text() if self._table.item(row, 6) else "",
            }
            test_cases.append(case)
        
        # 获取接口配置
        base_url = self._base_url_edit.text().strip()
        api_path = self._api_path_edit.text().strip()
        common_headers = _common_headers_store
        api_document = self._doc_input.toPlainText()
        
        # 生成默认名称
        default_name = f"{api_path.split('/')[-1] if api_path else '用例'}_{datetime.now().strftime('%m%d_%H%M')}"
        
        # 弹出输入对话框
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(
            self,
            "保存到历史记录",
            "请输入历史记录名称:",
            QLineEdit.EchoMode.Normal,
            default_name
        )
        
        if not ok or not name.strip():
            return
        
        # 询问是否收藏
        reply = QMessageBox.question(
            self,
            "收藏设置",
            "是否将此记录标记为收藏？\n（收藏记录不受50条限制）",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        is_favorite = reply == QMessageBox.StandardButton.Yes
        
        # 创建历史记录
        history = TestCaseHistory(
            name=name.strip(),
            description=f"接口: {api_path}",
            test_cases=test_cases,
            base_url=base_url,
            api_path=api_path,
            common_headers=common_headers,
            api_document=api_document,
            is_favorite=is_favorite,
        )
        
        try:
            history.save()
            
            msg = f"历史记录保存成功！\n名称: {history.name}\n用例数: {len(test_cases)}"
            if is_favorite:
                msg += "\n已标记为收藏"
            
            QMessageBox.information(self, "成功", msg)
            _logger.info(f"保存历史记录成功: {history.name}, 用例数: {len(test_cases)}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")
            _logger.error(f"保存历史记录失败: {str(e)}")
    
    def _show_history_dialog(self) -> None:
        """显示历史记录对话框"""
        dialog = TestCaseHistoryDialog(
            on_load_history=self._load_history_to_table,
            parent=self
        )
        dialog.exec()

    def _show_dimension_config(self) -> None:
        """显示测试维度配置对话框"""
        dialog = DimensionConfigDialog(parent=self)
        dialog.exec()
    
    def _load_history_to_table(self, history: TestCaseHistory) -> None:
        """将历史记录加载到表格"""
        self._table.setRowCount(0)

        if history.base_url or history.api_path:
            self._base_url_edit.setText(history.base_url or "")
            self._api_path_edit.setText(history.api_path or "")
            self._base_url_display.setText(history.base_url or "")
            self._api_path_display.setText(history.api_path or "")
            self._api_info_bar.setVisible(True)
        else:
            self._api_info_bar.setVisible(False)

        if history.common_headers:
            set_common_headers(history.common_headers)
        if history.api_document:
            self._doc_input.setPlainText(history.api_document)
            self._update_doc_status()
        
        # 加载用例到表格
        test_cases = history.test_cases if history.test_cases else []
        self._table.setRowCount(len(test_cases))
        
        for row, case_data in enumerate(test_cases):
            # 复用 _load_cases_to_table 的逻辑
            checkbox = QCheckBox()
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 0px;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                    border-radius: 3px;
                    border: 2px solid #bdbdbd;
                }
                QCheckBox::indicator:checked {
                    background-color: #2196f3;
                    border-color: #2196f3;
                }
                QCheckBox::indicator:hover {
                    border-color: #2196f3;
                }
            """)
            self._table.setCellWidget(row, 0, checkbox)
            
            self._table.setItem(row, 1, QTableWidgetItem(case_data.get("name", "")))
            self._table.setItem(row, 2, QTableWidgetItem(case_data.get("method", "GET")))
            self._table.setItem(row, 3, QTableWidgetItem(str(case_data.get("params", ""))))
            self._table.setItem(row, 4, QTableWidgetItem(str(case_data.get("body", ""))))
            
            status_combo = QComboBox()
            for code, desc in HTTP_STATUS_CODES.items():
                status_combo.addItem(desc, code)
            expected_status = case_data.get("expected_status", 200)
            status_combo.setCurrentText(HTTP_STATUS_CODES.get(str(expected_status), str(expected_status)))
            status_combo.currentIndexChanged.connect(lambda _, r=row: self._on_status_changed(r))
            status_combo.setStyleSheet("""
                QComboBox {
                    background-color: #fafafa;
                    border: 1px solid #e0e0e0;
                    border-radius: 4px;
                    padding: 4px 8px;
                    font-size: 12px;
                    min-width: 140px;
                }
                QComboBox:hover {
                    border-color: #bdbdbd;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox QAbstractItemView {
                    background-color: #ffffff;
                    border: 1px solid #e0e0e0;
                    selection-background-color: #e3f2fd;
                    min-width: 200px;
                }
            """)
            self._table.setCellWidget(row, 5, status_combo)
            
            self._table.setItem(row, 6, QTableWidgetItem(str(case_data.get("assertions", ""))))
        
        self._select_all_cb.setChecked(False)
        self._update_count()
        self._save_to_store()
        
        QMessageBox.information(
            self,
            "加载成功",
            f"已加载历史记录 '{history.name}'\n共 {len(test_cases)} 个测试用例"
        )
        _logger.info(f"加载历史记录到表格: {history.name}, 用例数: {len(test_cases)}")
